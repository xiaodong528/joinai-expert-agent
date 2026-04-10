#!/usr/bin/env python3
"""
spreadsheet_reader.py — Read .xlsx/.xls files into a workbook-render friendly JSON structure.

Supports:
  --input <file> --list-sheets
  --input <file> --sheet <name> --output <path.json>
  --input <file> --all-sheets --output-dir <dir>
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any


UNIT_KEYWORDS = frozenset({
    "元", "万元", "工日", "台班", "km", "m", "%", "个", "条", "套",
    "根", "组", "对", "米", "公里", "人", "项", "处", "块", "芯公里",
})
ROMAN_NUMERALS = frozenset({
    "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X",
})
NOISE_TEXTS = frozenset({
    "手动修改",
})
METADATA_PREFIXES = (
    "建设项目名称",
    "单项工程名称",
    "建设单位名称",
    "表格编号",
    "第 全 页",
)
LABEL_CONTEXT_KEYWORDS = ("费用名称", "名称", "项目名称", "规格程式")
AMOUNT_ROLE_KEYWORDS = (
    ("除税价", "pre_tax"),
    ("增值税", "tax"),
    ("含税价", "tax_inclusive"),
)


def col_letter(n: int) -> str:
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def cell_ref(row: int, col: int) -> str:
    return f"{col_letter(col)}{row}"


def infer_data_type(value: Any) -> str:
    if value in (None, ""):
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    return "string"


def normalize_cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def normalize_display_value(value: Any) -> Any:
    normalized = normalize_cell_value(value)
    if normalized is None:
        return ""
    return normalized


def build_merged_region(
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    value: Any,
) -> dict[str, Any]:
    return {
        "range_ref": f"{cell_ref(start_row, start_col)}:{cell_ref(end_row, end_col)}",
        "anchor_cell_ref": cell_ref(start_row, start_col),
        "start_row": start_row,
        "end_row": end_row,
        "start_col": start_col,
        "end_col": end_col,
        "row_span": end_row - start_row + 1,
        "col_span": end_col - start_col + 1,
        "value": normalize_display_value(value),
    }


def build_merge_lookups(merged_regions: list[dict[str, Any]]) -> tuple[dict[tuple[int, int], Any], dict[tuple[int, int], dict[str, Any]]]:
    merged_value_lookup: dict[tuple[int, int], Any] = {}
    merged_meta_lookup: dict[tuple[int, int], dict[str, Any]] = {}

    for region in merged_regions:
        for row_num in range(region["start_row"], region["end_row"] + 1):
            for col_num in range(region["start_col"], region["end_col"] + 1):
                merged_value_lookup[(row_num, col_num)] = region["value"]
                merged_meta_lookup[(row_num, col_num)] = {
                    "range_ref": region["range_ref"],
                    "anchor_cell_ref": region["anchor_cell_ref"],
                    "role": "anchor" if (row_num, col_num) == (region["start_row"], region["start_col"]) else "covered",
                    "start_row": region["start_row"],
                    "end_row": region["end_row"],
                    "start_col": region["start_col"],
                    "end_col": region["end_col"],
                    "row_span": region["row_span"],
                    "col_span": region["col_span"],
                }

    return merged_value_lookup, merged_meta_lookup


def is_unit_like_text(text: str) -> bool:
    stripped = text.strip()
    return stripped in UNIT_KEYWORDS or stripped in ROMAN_NUMERALS


def is_noise_text(text: str) -> bool:
    return text.strip() in NOISE_TEXTS


def detect_unit_row(cells: list[dict[str, Any]]) -> bool:
    text_cells = [
        cell for cell in cells
        if isinstance(cell.get("display_value"), str) and cell["display_value"].strip()
    ]
    if not text_cells:
        return False
    unit_like_count = sum(1 for cell in text_cells if is_unit_like_text(cell["display_value"]))
    return unit_like_count >= len(text_cells) * 0.5


def compute_row_context(cells: list[dict[str, Any]], is_unit_row: bool) -> str:
    if is_unit_row:
        return ""
    for cell in cells:
        value = cell.get("display_value")
        if isinstance(value, str):
            stripped = value.strip()
            if stripped and not is_unit_like_text(stripped) and not is_noise_text(stripped):
                return stripped
    return ""


def is_effective_cell(cell: dict[str, Any]) -> bool:
    if str(cell.get("formula", "") or "").strip():
        return True
    display_value = cell.get("display_value")
    if isinstance(display_value, str):
        return bool(display_value.strip())
    return display_value not in (None, "")


def is_effective_merged_region(region: dict[str, Any]) -> bool:
    value = region.get("value")
    if isinstance(value, str):
        return bool(value.strip())
    return value not in (None, "")


def trim_to_effective_bounds(
    rows: list[dict[str, Any]],
    merged_regions: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], int, int, str]:
    effective_regions = [region for region in merged_regions if is_effective_merged_region(region)]
    max_row = 0
    max_col = 0

    for row in rows:
        for col_index, cell in enumerate(row["cells"], start=1):
            if is_effective_cell(cell):
                max_row = max(max_row, row["row_num"])
                max_col = max(max_col, col_index)

    for region in effective_regions:
        max_row = max(max_row, region["end_row"])
        max_col = max(max_col, region["end_col"])

    if max_row == 0 or max_col == 0:
        return [], 0, 0, "A1:A1"

    trimmed_rows: list[dict[str, Any]] = []
    for row in rows:
        if row["row_num"] > max_row:
            break
        trimmed_rows.append(
            {
                "row_num": row["row_num"],
                "cells": row["cells"][:max_col],
            }
        )

    return trimmed_rows, max_row, max_col, f"A1:{cell_ref(max_row, max_col)}"


def row_text_cells(row: dict[str, Any]) -> list[str]:
    texts: list[str] = []
    for cell in row["cells"]:
        value = cell.get("display_value")
        if isinstance(value, str):
            stripped = value.strip()
            if stripped:
                texts.append(stripped)
    return texts


def is_title_like_row(row: dict[str, Any]) -> bool:
    texts = [text for text in row_text_cells(row) if not is_noise_text(text) and not is_unit_like_text(text)]
    if len(texts) < 3:
        return False
    return len(set(texts)) == 1


def row_has_numeric_or_formula(row: dict[str, Any]) -> bool:
    for cell in row["cells"]:
        if str(cell.get("formula", "") or "").strip():
            return True
        value = cell.get("display_value")
        if isinstance(value, (int, float)):
            return True
    return False


def header_candidate_rows(rows: list[dict[str, Any]]) -> list[int]:
    first_data_row_index = None
    for index, row in enumerate(rows):
        texts = row_text_cells(row)
        if not texts:
            continue
        if row.get("is_unit_row"):
            continue
        if row_has_numeric_or_formula(row):
            first_data_row_index = index
            break

    limit = first_data_row_index if first_data_row_index is not None else min(len(rows), 4)
    candidates: list[int] = []
    for index in range(limit):
        row = rows[index]
        texts = row_text_cells(row)
        if not texts:
            continue
        if row.get("is_unit_row"):
            continue
        if is_title_like_row(row):
            continue
        candidates.append(index)
    return candidates


def resolve_header_text(cell: dict[str, Any]) -> str:
    value = cell.get("display_value")
    if not isinstance(value, str):
        return ""
    stripped = value.strip()
    if not stripped or is_unit_like_text(stripped) or is_noise_text(stripped):
        return ""
    return stripped


def is_metadata_text(text: str) -> bool:
    stripped = text.strip()
    if not stripped:
        return True
    return any(stripped.startswith(prefix) for prefix in METADATA_PREFIXES)


def normalize_col_context(col_context: str) -> str:
    parts = [part.strip() for part in col_context.split(" / ") if part.strip()]
    normalized: list[str] = []
    for part in parts:
        if is_metadata_text(part):
            continue
        if normalized and normalized[-1] == part:
            continue
        normalized.append(part)
    return " / ".join(normalized)


def is_code_like_text(text: str) -> bool:
    stripped = text.strip()
    if not stripped:
        return False
    return bool(re.fullmatch(r"[A-Za-z]+[-\w（）()]*\d+(?:\.\d+)?", stripped))


def choose_row_business_label(cells: list[dict[str, Any]], row_context: str) -> str:
    prioritized: list[str] = []
    fallback: list[str] = []
    for cell in cells:
        value = cell.get("display_value")
        if not isinstance(value, str):
            continue
        stripped = value.strip()
        if (
            not stripped
            or is_unit_like_text(stripped)
            or is_noise_text(stripped)
            or is_metadata_text(stripped)
            or is_code_like_text(stripped)
        ):
            continue
        col_context = str(cell.get("col_context", "") or "")
        if any(keyword in col_context for keyword in LABEL_CONTEXT_KEYWORDS):
            prioritized.append(stripped)
        fallback.append(stripped)
    if prioritized:
        return prioritized[0]
    if row_context and not is_code_like_text(row_context):
        return row_context
    return fallback[0] if fallback else row_context


def infer_amount_role(col_context: str) -> str | None:
    for keyword, role in AMOUNT_ROLE_KEYWORDS:
        if keyword in col_context:
            return role
    return None


def compute_col_contexts(rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return []
    candidates = header_candidate_rows(rows)
    if not candidates:
        return [""] * len(rows[0]["cells"])

    col_count = max(len(row["cells"]) for row in rows)
    contexts: list[str] = []
    for col_index in range(col_count):
        parts: list[str] = []
        for row_index in candidates:
            row = rows[row_index]
            if col_index >= len(row["cells"]):
                continue
            text = resolve_header_text(row["cells"][col_index])
            if text and (not parts or parts[-1] != text):
                parts.append(text)
        contexts.append(" / ".join(parts))
    return contexts


def enrich_contexts(rows: list[dict[str, Any]]) -> None:
    for row in rows:
        row["is_unit_row"] = detect_unit_row(row["cells"])

    col_contexts = compute_col_contexts(rows)
    for row_index, row in enumerate(rows):
        row_context = compute_row_context(row["cells"], row["is_unit_row"])
        for col_index, cell in enumerate(row["cells"]):
            cell["row_context"] = row_context
            cell["col_context"] = normalize_col_context(
                col_contexts[col_index] if col_index < len(col_contexts) else ""
            )
        row_business_label = choose_row_business_label(row["cells"], row_context)
        for cell in row["cells"]:
            cell["row_business_label"] = row_business_label
            cell["amount_role"] = infer_amount_role(str(cell.get("col_context", "") or ""))


def infer_formula_annotations(rows: list[dict[str, Any]]) -> None:
    for row in rows:
        for cell in row["cells"]:
            if "formula_annotation" not in cell:
                cell["formula_annotation"] = None
            if "rule_annotations" not in cell:
                cell["rule_annotations"] = []
            formula = str(cell.get("formula", "") or "").strip()
            if formula:
                cell["formula_annotation"] = {
                    "expression": f"{cell['cell_ref']}={formula.lstrip('=')}",
                    "source": "native",
                    "confidence": "high",
                }


def build_sheet_payload(
    sheet_name: str,
    rows: list[dict[str, Any]],
    merged_regions: list[dict[str, Any]],
) -> dict[str, Any]:
    effective_regions = [region for region in merged_regions if is_effective_merged_region(region)]
    trimmed_rows, row_count, col_count, dimensions = trim_to_effective_bounds(rows, effective_regions)
    enrich_contexts(trimmed_rows)
    infer_formula_annotations(trimmed_rows)
    return {
        "sheet_name": sheet_name,
        "dimensions": dimensions,
        "row_count": row_count,
        "col_count": col_count,
        "rows": trimmed_rows,
        "merged_cells": [region["range_ref"] for region in effective_regions],
        "merged_regions": effective_regions,
    }


def read_xlsx(filepath: str) -> dict[str, dict[str, Any]]:
    import openpyxl

    value_wb = openpyxl.load_workbook(filepath, data_only=True)
    formula_wb = openpyxl.load_workbook(filepath, data_only=False)
    result: dict[str, dict[str, Any]] = {}

    try:
        for formula_ws in formula_wb.worksheets:
            value_ws = value_wb[formula_ws.title]
            sheet_name = formula_ws.title
            merged_regions: list[dict[str, Any]] = []
            for merged_range in formula_ws.merged_cells.ranges:
                anchor_value = value_ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                merged_regions.append(
                    build_merged_region(
                        start_row=merged_range.min_row,
                        end_row=merged_range.max_row,
                        start_col=merged_range.min_col,
                        end_col=merged_range.max_col,
                        value=anchor_value,
                    )
                )
            merged_value_lookup, merged_meta_lookup = build_merge_lookups(merged_regions)

            min_row = formula_ws.min_row or 1
            max_row = formula_ws.max_row or 1
            min_col = formula_ws.min_column or 1
            max_col = formula_ws.max_column or 1

            rows: list[dict[str, Any]] = []
            for row_num in range(min_row, max_row + 1):
                cells: list[dict[str, Any]] = []
                for col_num in range(min_col, max_col + 1):
                    ref = cell_ref(row_num, col_num)
                    lookup_key = (row_num, col_num)
                    raw_value = merged_value_lookup.get(lookup_key, value_ws.cell(row=row_num, column=col_num).value)
                    display_value = normalize_display_value(raw_value)
                    formula_value = formula_ws.cell(row=row_num, column=col_num).value
                    formula = formula_value if isinstance(formula_value, str) and formula_value.startswith("=") else ""
                    if isinstance(display_value, float):
                        display_value = round(display_value, 10)
                    value = display_value
                    cell_payload: dict[str, Any] = {
                        "cell_ref": ref,
                        "value": value,
                        "display_value": display_value,
                        "data_type": infer_data_type(display_value),
                        "formula": formula,
                        "formula_annotation": None,
                        "rule_annotations": [],
                    }
                    merge_metadata = merged_meta_lookup.get(lookup_key)
                    if merge_metadata:
                        cell_payload["merge"] = merge_metadata
                    cells.append(cell_payload)
                rows.append({"row_num": row_num, "cells": cells})

            result[sheet_name] = build_sheet_payload(
                sheet_name=sheet_name,
                rows=rows,
                merged_regions=merged_regions,
            )
    finally:
        value_wb.close()
        formula_wb.close()

    return result


def _xlrd_cell_value(ws: Any, row: int, col: int) -> Any:
    import xlrd

    cell = ws.cell(row, col)
    ctype = cell.ctype
    if ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if ctype == xlrd.XL_CELL_TEXT:
        return cell.value
    if ctype == xlrd.XL_CELL_NUMBER:
        return cell.value
    if ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(cell.value, ws.book.datemode).isoformat()
        except Exception:
            return cell.value
    if ctype == xlrd.XL_CELL_ERROR:
        return None
    return cell.value


def read_xls(filepath: str) -> dict[str, dict[str, Any]]:
    import xlrd

    workbook = xlrd.open_workbook(filepath, formatting_info=True)
    result: dict[str, dict[str, Any]] = {}

    for sheet_index in range(workbook.nsheets):
        worksheet = workbook.sheet_by_index(sheet_index)
        sheet_name = worksheet.name
        merged_regions: list[dict[str, Any]] = []
        for row_lo, row_hi, col_lo, col_hi in worksheet.merged_cells:
            anchor_value = _xlrd_cell_value(worksheet, row_lo, col_lo)
            merged_regions.append(
                build_merged_region(
                    start_row=row_lo + 1,
                    end_row=row_hi,
                    start_col=col_lo + 1,
                    end_col=col_hi,
                    value=anchor_value,
                )
            )
        merged_value_lookup, merged_meta_lookup = build_merge_lookups(merged_regions)

        rows: list[dict[str, Any]] = []
        for row_num in range(worksheet.nrows):
            cells: list[dict[str, Any]] = []
            for col_num in range(worksheet.ncols):
                ref = cell_ref(row_num + 1, col_num + 1)
                lookup_key = (row_num + 1, col_num + 1)
                raw_value = merged_value_lookup.get(lookup_key, _xlrd_cell_value(worksheet, row_num, col_num))
                display_value = normalize_display_value(raw_value)
                if isinstance(display_value, float):
                    display_value = round(display_value, 10)
                cell_payload: dict[str, Any] = {
                    "cell_ref": ref,
                    "value": display_value,
                    "display_value": display_value,
                    "data_type": infer_data_type(display_value),
                    "formula": "",
                    "formula_annotation": None,
                    "rule_annotations": [],
                }
                merge_metadata = merged_meta_lookup.get(lookup_key)
                if merge_metadata:
                    cell_payload["merge"] = merge_metadata
                cells.append(cell_payload)
            rows.append({"row_num": row_num + 1, "cells": cells})

        row_count = worksheet.nrows
        col_count = worksheet.ncols
        result[sheet_name] = build_sheet_payload(
            sheet_name=sheet_name,
            rows=rows,
            merged_regions=merged_regions,
        )

    return result


def read_file(filepath: str) -> dict[str, dict[str, Any]]:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xlsx":
        return read_xlsx(filepath)
    if ext == ".xls":
        return read_xls(filepath)
    raise ValueError(f"Unsupported file format: {ext!r}. Use .xlsx or .xls")


def write_json(data: dict[str, Any], output_path: str) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def safe_filename(name: str) -> str:
    safe = "".join(char if char.isalnum() or char in "-_. " else "_" for char in name)
    return safe.strip()


def main() -> int:
    parser = argparse.ArgumentParser(description="Read .xlsx/.xls spreadsheets into workbook-render friendly JSON.")
    parser.add_argument("--input", required=True, help="Path to .xlsx or .xls file")
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--list-sheets", action="store_true", help="Print all sheet names")
    mode.add_argument("--sheet", metavar="NAME", help="Extract one sheet by name")
    mode.add_argument("--all-sheets", action="store_true", help="Extract all sheets")
    parser.add_argument("--output", metavar="PATH", help="Output JSON path (for --sheet)")
    parser.add_argument("--output-dir", metavar="DIR", help="Output directory (for --all-sheets)")
    args = parser.parse_args()

    if not os.path.isfile(args.input):
        print(f"Error: input file not found: {args.input}", file=sys.stderr)
        return 1

    try:
        sheets = read_file(args.input)
        if args.list_sheets:
            for name in sheets:
                print(name)
            return 0

        if args.sheet:
            if not args.output:
                print("Error: --output is required with --sheet", file=sys.stderr)
                return 1
            if args.sheet not in sheets:
                available = ", ".join(repr(name) for name in sheets)
                print(f"Error: sheet {args.sheet!r} not found. Available: {available}", file=sys.stderr)
                return 1
            write_json(sheets[args.sheet], args.output)
            print(f"Wrote {args.output}")
            return 0

        if args.all_sheets:
            if not args.output_dir:
                print("Error: --output-dir is required with --all-sheets", file=sys.stderr)
                return 1
            os.makedirs(args.output_dir, exist_ok=True)
            for name, data in sheets.items():
                output_path = os.path.join(args.output_dir, f"{safe_filename(name)}.json")
                write_json(data, output_path)
                print(f"Wrote {output_path}")
            return 0
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
