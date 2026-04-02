import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl
import yaml


ROOT = Path(__file__).resolve().parents[4]
SCRIPT = ROOT / ".opencode/skills/construction-audit-s2-workbook-render/scripts/run_workbook_render.py"
SAMPLE_XLS = ROOT / "train/预算-表一（451定额度折前）/东海县海陵家苑三网小区新建工程-预算（表一451定额度折前有错）.xls"
RULE_DOC = ROOT / "examples/rules-docx/家客预算审核知识库11.9.docx"


class RunWorkbookRenderTests(unittest.TestCase):
    def _write_config(self, path: Path, config: dict) -> None:
        path.write_text(yaml.safe_dump(config, allow_unicode=True, sort_keys=False), encoding="utf-8")

    def _run(self, config_path: Path) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [sys.executable, str(SCRIPT), "--config", str(config_path)],
            capture_output=True,
            text=True,
            cwd=ROOT,
        )

    def test_renders_workbook_and_sheet_jsons_from_minimal_config(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            workbook_path = tmpdir_path / "spreadsheet.xlsx"
            config_path = tmpdir_path / "audit-config.yaml"
            output_dir = tmpdir_path / "output"

            workbook = openpyxl.Workbook()
            workbook.active.title = "工程信息"
            workbook.active["A1"] = "项目名称"
            workbook.active["B1"] = "东海县海陵家苑三网小区新建工程"
            workbook.create_sheet("表一")
            workbook["表一"]["A1"] = "费用名称"
            workbook["表一"]["B1"] = "金额"
            workbook["表一"]["A2"] = "人工费"
            workbook["表一"]["B2"] = 10
            workbook.save(workbook_path)

            self._write_config(
                config_path,
                {
                    "audit_id": "AUDIT-WB-001",
                    "spreadsheet": {
                        "path": str(workbook_path),
                        "sheets": ["工程信息", "表一"],
                    },
                    "output_dir": str(output_dir),
                },
            )

            result = self._run(config_path)

            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertIn("audit_id=AUDIT-WB-001", result.stdout)
            self.assertIn(f"input_workbook={workbook_path.resolve()}", result.stdout)
            self.assertIn("target_sheets_count=2", result.stdout)
            workbook_md = output_dir / "workbook.md"
            self.assertTrue(workbook_md.exists())
            self.assertTrue((output_dir / "sheets" / "工程信息.json").exists())
            self.assertTrue((output_dir / "sheets" / "表一.json").exists())
            content = workbook_md.read_text(encoding="utf-8")
            self.assertIn("## Sheet: 工程信息", content)
            self.assertIn("## Sheet: 表一", content)
            self.assertIn("### Sheet View", content)

    def test_fails_when_spreadsheet_path_missing(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            config_path = tmpdir_path / "audit-config.yaml"
            self._write_config(
                config_path,
                {
                    "audit_id": "AUDIT-WB-002",
                    "spreadsheet": {
                        "sheets": ["工程信息"],
                    },
                    "output_dir": str(tmpdir_path / "output"),
                },
            )

            result = self._run(config_path)

            self.assertNotEqual(result.returncode, 0)
            self.assertIn("spreadsheet.path", result.stderr)

    def test_fails_when_target_sheets_missing(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            workbook_path = tmpdir_path / "spreadsheet.xlsx"
            config_path = tmpdir_path / "audit-config.yaml"

            workbook = openpyxl.Workbook()
            workbook.active.title = "工程信息"
            workbook.save(workbook_path)

            self._write_config(
                config_path,
                {
                    "audit_id": "AUDIT-WB-003",
                    "spreadsheet": {
                        "path": str(workbook_path),
                    },
                    "output_dir": str(tmpdir_path / "output"),
                },
            )

            result = self._run(config_path)

            self.assertNotEqual(result.returncode, 0)
            self.assertIn("spreadsheet.sheets", result.stderr)

    def test_real_xls_regression_uses_trimmed_business_range(self):
        if not SAMPLE_XLS.exists():
            self.skipTest(f"Sample workbook not found: {SAMPLE_XLS}")

        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            config_path = tmpdir_path / "audit-config.yaml"
            output_dir = tmpdir_path / "output"
            self._write_config(
                config_path,
                {
                    "audit_id": "AUDIT-WB-REAL-001",
                    "rule_document": {
                        "path": str(RULE_DOC),
                        "markdown_path": str(output_dir / "rule_doc.md"),
                    },
                    "spreadsheet": {
                        "path": str(SAMPLE_XLS),
                        "sheets": ["工程信息", "表一（451定额折前）"],
                    },
                    "output_dir": str(output_dir),
                },
            )

            result = self._run(config_path)

            self.assertEqual(result.returncode, 0, result.stderr)
            workbook_md = output_dir / "workbook.md"
            self.assertTrue(workbook_md.exists())
            table_payload = json.loads((output_dir / "sheets" / "表一_451定额折前_.json").read_text(encoding="utf-8"))
            self.assertEqual(table_payload["dimensions"], "A1:M15")
            content = workbook_md.read_text(encoding="utf-8")
            self.assertIn("used_range: A1:M15", content)
            row7_label = next(cell for row in table_payload["rows"] if row["row_num"] == 7 for cell in row["cells"] if cell["cell_ref"] == "C7")
            row7_value = next(cell for row in table_payload["rows"] if row["row_num"] == 7 for cell in row["cells"] if cell["cell_ref"] == "G7")
            row7_pre_tax = next(cell for row in table_payload["rows"] if row["row_num"] == 7 for cell in row["cells"] if cell["cell_ref"] == "J7")
            row7_tax = next(cell for row in table_payload["rows"] if row["row_num"] == 7 for cell in row["cells"] if cell["cell_ref"] == "K7")
            row7_total = next(cell for row in table_payload["rows"] if row["row_num"] == 7 for cell in row["cells"] if cell["cell_ref"] == "L7")
            row8_pre_tax = next(cell for row in table_payload["rows"] if row["row_num"] == 8 for cell in row["cells"] if cell["cell_ref"] == "J8")
            row8_tax = next(cell for row in table_payload["rows"] if row["row_num"] == 8 for cell in row["cells"] if cell["cell_ref"] == "K8")
            row8_total = next(cell for row in table_payload["rows"] if row["row_num"] == 8 for cell in row["cells"] if cell["cell_ref"] == "L8")
            row9_pre_tax = next(cell for row in table_payload["rows"] if row["row_num"] == 9 for cell in row["cells"] if cell["cell_ref"] == "J9")
            row9_tax = next(cell for row in table_payload["rows"] if row["row_num"] == 9 for cell in row["cells"] if cell["cell_ref"] == "K9")
            row9_total = next(cell for row in table_payload["rows"] if row["row_num"] == 9 for cell in row["cells"] if cell["cell_ref"] == "L9")
            row12_label = next(cell for row in table_payload["rows"] if row["row_num"] == 12 for cell in row["cells"] if cell["cell_ref"] == "C12")
            row12_value = next(cell for row in table_payload["rows"] if row["row_num"] == 12 for cell in row["cells"] if cell["cell_ref"] == "J12")
            row13_label = next(cell for row in table_payload["rows"] if row["row_num"] == 13 for cell in row["cells"] if cell["cell_ref"] == "C13")
            row13_value = next(cell for row in table_payload["rows"] if row["row_num"] == 13 for cell in row["cells"] if cell["cell_ref"] == "I13")
            self.assertTrue(row7_label["rule_annotations"])
            self.assertEqual(len(row7_label["rule_annotations"]), 1)
            self.assertEqual(row7_label["rule_annotations"][0]["role"], "target")
            self.assertEqual(row7_label["rule_annotations"][0]["source_anchor"], "表一（451定额折前）审核规则")
            self.assertEqual(row7_label["rule_annotations"][0]["source_row_index"], 1)
            self.assertEqual({(ann["role"], ann["source_row_index"]) for ann in row7_value["rule_annotations"]}, {("target", 1), ("operand", 6), ("operand", 7)})
            self.assertEqual(len(row12_label["rule_annotations"]), 1)
            self.assertEqual({(ann["role"], ann["source_row_index"]) for ann in row12_label["rule_annotations"]}, {("target", 6)})
            self.assertEqual({(ann["role"], ann["source_row_index"]) for ann in row12_value["rule_annotations"]}, {("target", 6), ("operand", 8)})
            self.assertEqual({(ann["role"], ann["source_row_index"]) for ann in row13_label["rule_annotations"]}, {("target", 7)})
            self.assertEqual({(ann["role"], ann["source_row_index"]) for ann in row13_value["rule_annotations"]}, {("target", 7), ("operand", 8)})
            self.assertEqual(row7_pre_tax["row_business_label"], "建筑安装工程费")
            self.assertEqual(row7_total["row_business_label"], "建筑安装工程费")
            self.assertEqual(row8_pre_tax["row_business_label"], "需要安装的设备费")
            self.assertEqual(row8_total["row_business_label"], "需要安装的设备费")
            self.assertEqual(row9_pre_tax["row_business_label"], "工程建设其他费")
            self.assertEqual(row9_total["row_business_label"], "工程建设其他费")
            self.assertEqual(row7_pre_tax["amount_role"], "pre_tax")
            self.assertEqual(row7_tax["amount_role"], "tax")
            self.assertEqual(row7_total["amount_role"], "tax_inclusive")
            self.assertEqual(row8_pre_tax["amount_role"], "pre_tax")
            self.assertEqual(row8_tax["amount_role"], "tax")
            self.assertEqual(row8_total["amount_role"], "tax_inclusive")
            self.assertEqual(row9_pre_tax["amount_role"], "pre_tax")
            self.assertEqual(row9_tax["amount_role"], "tax")
            self.assertEqual(row9_total["amount_role"], "tax_inclusive")
            self.assertEqual(row7_tax["col_context"], "总价值 / 增值税(元)")
            self.assertEqual(row7_total["col_context"], "总价值 / 含税价(元)")
            self.assertIsNone(row7_total["formula_annotation"])
            self.assertNotIn("(L7=J7+K7)", content)
            self.assertNotIn("[RULE ", content)
            self.assertNotIn("安全生产费计算表（表二折前）建筑安装工程费所对应的合计（元）", content)
            self.assertIn("| 7 | 1.0 | TXL-2 | 建筑安装工程费 |", content)


if __name__ == "__main__":
    unittest.main()
